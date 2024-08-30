#  ____       _____           _
# |  _ \ _   |_   _|__   ___ | |___ ____
# | |_) | | | || |/ _ \ / _ \| / __|_  /
# |  __/| |_| || | (_) | (_) | \__ \/ /
# |_|    \__, ||_|\___/ \___/|_|___/___|
#        |___/

# Copyright (c) 2024 Sidney Zhang <zly@lyzhang.me>
# PyToolsz is licensed under Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#          http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.

# 说明
# 这是一个把pandas/polars的DataFrame数据保存成一个具有格式的excel文件。
# 例如把计算好的结算数据写到excel文件中，依照格式直接形成结算表。

from pathlib import Path
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.worksheet import cell_range
import itertools
import string
from openpyxl.utils.cell import get_column_letter, cols_from_range, coordinate_from_string, column_index_from_string
from collections import ChainMap
from collections.abc import Iterable

class saveExcel(object):
    """
    保存DataFrame到excel文件。
    """
    def __init__(self, filename:str|Path, title:str|None = None, sheetname:str = "Sheet1",
                 startRow:int = 1, startColumn:int = 1) -> None:
        self.__filename = filename
        self.__wb = Workbook()
        self.__sheetname = sheetname
        self.__startRow = startRow
        self.__startColumn = startColumn
        self.__place = None
    def __enter__(self) -> any:
        return self.__wb.active
    def __exit__(self, exc_type, exc_value, traceback) -> None:
        self.__wb.save(self.__filename)